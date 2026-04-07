from __future__ import annotations

"""
Reports blueprint: Chart.js charts, CSV/XLSX export, JSON backup.
"""

import csv
import io
import json
from datetime import datetime, timezone
from typing import Any

from flask import Blueprint, Response, g, jsonify, render_template, request, stream_with_context
from flask_login import current_user, login_required

from .common import require_admin

reports_bp = Blueprint("reports", __name__)

_CHARTJS = "/static/chart.umd.min.js"


def _active_alerts() -> int:
    return len(g.inv.get_active_alerts())


# ─── main reports page ────────────────────────────────────────────────────────

@reports_bp.route("/reports")
@login_required
def reports_index():
    try:
        days = max(1, min(365, int(request.args.get("days", 30))))
    except (ValueError, TypeError):
        days = 30
    inv = g.inv

    usage = inv.chart_usage_by_material(days=days)
    jobs_per_day = inv.chart_jobs_per_day(days=days)
    low_spools = inv.spools_below_threshold()

    # Donut chart: usage by material
    mat_labels = json.dumps([r["material"] for r in usage])
    mat_data = json.dumps([float(r["used_g"]) for r in usage])
    mat_colors = json.dumps(_palette(len(usage)))

    # Bar chart: jobs per day per printer
    printers_seen: list[str] = []
    days_seen: list[str] = []
    for r in jobs_per_day:
        day = str(r["day"])[:10]
        pname = r["printer_name"]
        if day not in days_seen:
            days_seen.append(day)
        if pname not in printers_seen:
            printers_seen.append(pname)
    days_seen.sort()
    jobs_matrix: dict[str, dict[str, int]] = {p: {} for p in printers_seen}
    for r in jobs_per_day:
        jobs_matrix[r["printer_name"]][str(r["day"])[:10]] = int(r["count"])
    bar_labels = json.dumps(days_seen)
    bar_datasets = json.dumps([
        {
            "label": p,
            "data": [jobs_matrix[p].get(d, 0) for d in days_seen],
            "backgroundColor": _palette(len(printers_seen))[i % len(printers_seen)],
        }
        for i, p in enumerate(printers_seen)
    ])

    return render_template(
        "reports.html",
        title="Reports",
        active_nav="Reports",
        alert_count=_active_alerts(),
        days=days,
        low_spools=low_spools,
        mat_labels=mat_labels,
        mat_data=mat_data,
        mat_colors=mat_colors,
        bar_labels=bar_labels,
        bar_datasets=bar_datasets,
        chartjs_url=_CHARTJS,
    )


# ─── spool stock chart (per spool) ────────────────────────────────────────────

@reports_bp.route("/reports/spool/<int:spool_id>/chart")
@login_required
def spool_chart(spool_id: int):
    inv = g.inv
    s = inv.get_spool_dict(spool_id)
    if not s:
        return "Not found", 404
    history = inv.chart_stock_over_time(spool_id)
    timestamps = json.dumps([str(r["timestamp"])[:16].replace("T", " ") for r in history])
    values = json.dumps([float(r["new_remaining_g"]) for r in history])

    return render_template(
        "spool_chart.html",
        title=f"Chart: {s['name']}",
        active_nav="",
        alert_count=_active_alerts(),
        s=s,
        spool_id=spool_id,
        timestamps=timestamps,
        values=values,
        chartjs_url=_CHARTJS,
    )


# ─── CSV export (streaming) ───────────────────────────────────────────────────

# Column names for the spools CSV — matches what export_spools_iter() yields
_SPOOL_FIELDNAMES = [
    "id", "barcode_id", "name", "material", "color_hex", "brand",
    "total_weight_g", "remaining_g", "low_stock_threshold_g",
    "purchase_date", "purchase_price_cents", "notes",
    "is_archived", "created_by", "created_at", "updated_at",
]


@reports_bp.route("/reports/export/spools.csv")
@login_required
def export_spools_csv():
    def generate():
        output = io.StringIO()
        # Use SPOOL_FIELDNAMES if available; fall back to first-row keys
        first_row = None
        fieldnames = _SPOOL_FIELDNAMES

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        yield output.getvalue()
        output.truncate(0)
        output.seek(0)

        for row in g.inv.export_spools_iter():
            writer.writerow({k: str(row.get(k, "") if row.get(k) is not None else "") for k in fieldnames})
            yield output.getvalue()
            output.truncate(0)
            output.seek(0)

    return Response(
        stream_with_context(generate()),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=spools.csv"},
    )


@reports_bp.route("/reports/export/jobs.csv")
@login_required
def export_jobs_csv():
    rows = g.inv.export_jobs_dicts()
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for r in rows:
            writer.writerow({k: str(v) if v is not None else "" for k, v in r.items()})
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=jobs.csv"},
    )


@reports_bp.route("/reports/export/events.csv")
@login_required
def export_events_csv():
    rows = g.inv.export_stock_events_dicts()
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for r in rows:
            writer.writerow({k: str(v) if v is not None else "" for k, v in r.items()})
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=stock_events.csv"},
    )


# ─── XLSX export ──────────────────────────────────────────────────────────────

def _dicts_to_xlsx_sheet(ws: Any, rows: list[dict]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([str(r.get(h, "")) if r.get(h) is not None else "" for h in headers])


@reports_bp.route("/reports/export/spools.xlsx")
@login_required
def export_spools_xlsx():
    try:
        import openpyxl
    except ImportError:
        return "openpyxl not installed", 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spools"
    _dicts_to_xlsx_sheet(ws, g.inv.export_spools_dicts())

    ws2 = wb.create_sheet("Stock Events")
    _dicts_to_xlsx_sheet(ws2, g.inv.export_stock_events_dicts())

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=spools.xlsx"},
    )


@reports_bp.route("/reports/export/jobs.xlsx")
@login_required
def export_jobs_xlsx():
    try:
        import openpyxl
    except ImportError:
        return "openpyxl not installed", 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Print Jobs"
    _dicts_to_xlsx_sheet(ws, g.inv.export_jobs_dicts())

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=jobs.xlsx"},
    )


# ─── full backup ──────────────────────────────────────────────────────────────

@reports_bp.route("/reports/backup.json")
@login_required
@require_admin
def backup_json():
    inv = g.inv
    inv.record_audit("backup.download", user_id=current_user.id,
                     ip_address=request.remote_addr or "")

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "spools": inv.export_spools_dicts(),
        "print_jobs": inv.export_jobs_dicts(),
        "stock_events": inv.export_stock_events_dicts(),
    }

    import gzip
    raw = json.dumps(payload, default=str).encode()
    buf = io.BytesIO()
    with gzip.GzipFile(fileobj=buf, mode="wb") as gz:
        gz.write(raw)

    return Response(
        buf.getvalue(),
        mimetype="application/gzip",
        headers={"Content-Disposition": "attachment; filename=bambu_tracker_backup.json.gz"},
    )


# ─── API: chart data endpoints ────────────────────────────────────────────────

@reports_bp.route("/api/charts/usage")
@login_required
def api_chart_usage():
    try:
        days = max(1, min(365, int(request.args.get("days", 30))))
    except (ValueError, TypeError):
        days = 30
    return jsonify(g.inv.chart_usage_by_material(days=days))


@reports_bp.route("/api/charts/jobs")
@login_required
def api_chart_jobs():
    try:
        days = max(1, min(365, int(request.args.get("days", 30))))
    except (ValueError, TypeError):
        days = 30
    return jsonify(g.inv.chart_jobs_per_day(days=days))


# ─── helpers ──────────────────────────────────────────────────────────────────

def _palette(n: int) -> list[str]:
    base = [
        "#4f46e5", "#06b6d4", "#10b981", "#f59e0b",
        "#ef4444", "#8b5cf6", "#ec4899", "#14b8a6",
        "#f97316", "#6366f1",
    ]
    return [base[i % len(base)] for i in range(max(n, 1))]
